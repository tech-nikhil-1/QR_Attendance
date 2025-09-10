import cv2
import os
import time
import tkinter as tk
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ---------- Excel Setup ----------
def setup_excel():
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"attendance_{today}.xlsx"

    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["S.No", "ID", "Name", "Date", "In-Time", "Out-Time"])
        workbook.save(filename)

    return workbook, sheet, filename

workbook, sheet, filename = setup_excel()

# ---------- Popup Display ----------
def show_popup(message, duration=4000, countdown=False):
    root = tk.Tk()
    root.withdraw()  # hide main window
    popup = tk.Toplevel(root)
    popup.title("Attendance Log")

    label = tk.Label(popup, text=message, font=("Arial", 12), padx=20, pady=20, justify="left")
    label.pack()

    if countdown:
        seconds = duration // 1000
        def update_countdown(sec):
            if sec > 0:
                label.config(text=f"{message}\n\nNow place QR for scan in {sec} sec(s)")
                popup.after(1000, update_countdown, sec - 1)
            else:
                popup.destroy()
                root.destroy()

        update_countdown(seconds)
    else:
        popup.after(duration, popup.destroy)
        root.after(duration, root.destroy)

    root.mainloop()

# ---------- QR Scanner ----------
def scan_qr_code():
    cap = cv2.VideoCapture(0)
    detector = cv2.QRCodeDetector()

    while True:
        _, frame = cap.read()
        data, bbox, _ = detector.detectAndDecode(frame)

        if bbox is not None:
            bbox = bbox.astype(int)
            for i in range(len(bbox)):
                pt1 = tuple(bbox[i][0])
                pt2 = tuple(bbox[(i + 1) % len(bbox)][0])
                cv2.line(frame, pt1, pt2, (0, 255, 0), 2)

            if data:  # return QR data once detected
                cap.release()
                cv2.destroyAllWindows()
                return data

        cv2.imshow("QR Scanner - Press Q to Quit", frame)
        if cv2.waitKey(1) == ord("q"):
            cap.release()
            cv2.destroyAllWindows()
            return None  # exit scanning loop

# ---------- Save Attendance ----------
def save_attendance(qr_data):
    global workbook, sheet, filename

    now = datetime.now()
    date, current_time = now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")

    # Parse QR data
    name, id_no = "Unknown", "Unknown"
    try:
        for part in qr_data.split("\n"):
            if "Name:" in part:
                name = part.split("Name:")[1].strip()
            elif "Id.No:" in part:
                id_no = part.split("Id.No:")[1].strip()
    except:
        pass

    # Check if last record for this ID today has empty Out-Time
    last_row = None
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[1].value == id_no and row[3].value == date:
            last_row = row

    if last_row and last_row[5].value == "":  
        # Update Out-Time
        last_row[5].value = current_time
        message = f"[INFO] {name} ({id_no}) -> Out-Time at {current_time}"
    else:
        # Create new In-Time row
        s_no = sheet.max_row  # auto serial number
        sheet.append([s_no, id_no, name, date, current_time, ""])
        message = f"[INFO] {name} ({id_no}) -> In-Time at {current_time}"

    workbook.save(filename)
    print(message)

    # Show popup log for scan result (with countdown)
    show_popup(message, duration=4000, countdown=True)

# ---------- Main Loop ----------
print("[INFO] Starting QR Attendance System")
while True:
    qr_data = scan_qr_code()
    if qr_data:
        save_attendance(qr_data)
    else:
        print("[INFO] Scanner stopped by user.")
        break
